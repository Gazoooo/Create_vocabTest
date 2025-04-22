from openpyxl import load_workbook
import tools as t

workbook = load_workbook(t.find_path("sources/Cursus_A.xlsx"))
sheet = workbook.active


if (sheet["A"][0].value != "Lektion" or sheet["B"][0].value != "lat" or sheet["C"][0].value != "art"):
    print("Datenbank wurde nicht erstellt. Bitte folgende Reihenfolge der Tabelle beachten:\nLektion - lat - art")
else:
    for value in sheet.iter_rows(min_col=1,max_col=4,values_only=True):
        if value[0] == "Lektion":
            pass
        else:
            lektion = value[0]
            vokabel = value[1]
            wortart = value[2]
            zusatz = value[3]
            print(zusatz)


