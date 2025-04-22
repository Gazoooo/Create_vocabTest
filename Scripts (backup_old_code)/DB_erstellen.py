from openpyxl import load_workbook
import sqlite3
import tools as t
import os
from PIL import Image
from tkinter import Tk, filedialog

root = Tk()
root.withdraw()
file = filedialog.askopenfilename(title="Bitte Datei auswählen", initialdir=t.find_path("sources"),filetypes=(('Excel files', '*.xlsx'), ('All files', '*.*')))

workbook = load_workbook(file)
sheet = workbook.active


if (file != "") and (sheet["A"][0].value != "Lektion" or sheet["B"][0].value != "lat" or sheet["C"][0].value != "art"):
    print("Datenbank wurde nicht erstellt. Bitte folgende Reihenfolge der Tabelle beachten:\nLektion - lat - art")
else:
    name_of_new_DB = f'{file[:file.rfind(".")]}.db'
    if os.path.isfile(t.find_path(name_of_new_DB)):
        os.remove(t.find_path(name_of_new_DB))
    verbindung = sqlite3.connect(name_of_new_DB)
    zeiger = verbindung.cursor()
    l = []
    for i in sheet["A"]:
        l.append(i.value)
    l.pop(0)

    for i in range(1, max(l)+1):
        zeiger.execute(f'CREATE TABLE "Vokabeln_Lektion_{i}" ("Latein" TEXT, "Wortart" TEXT, PRIMARY KEY("Latein") )')

    for value in sheet.iter_rows(min_col=1,max_col=3,values_only=True):
        if value[0] == "Lektion":
            pass
        else:
            lektion = value[0]
            vokabel = value[1]
            wortart = value[2]
            zeiger.execute(f'INSERT INTO Vokabeln_Lektion_{lektion} VALUES ("{vokabel}", "{wortart}")')
            verbindung.commit()

    verbindung.close()

