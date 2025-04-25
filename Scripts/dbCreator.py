from openpyxl import load_workbook
import sqlite3
import tools as t
import os
from tkinter import Tk, filedialog

#Class to creates Sqllite Databases from given Excel worksheets, started separately by "new_DB.bat"

class dbCreator:
    """
    This class creates an SQLite database from an Excel file with vocabulary data.
    It processes an Excel worksheet containing lesson vocabulary and creates a database
    with tables for each lesson.
    """
    
    def __init__(self):
        """
        Initializes the dbCreator class by asking the user to select an Excel file
        and loading the workbook and the active sheet.

        Prompts the user to select an Excel file containing vocabulary data, 
        then loads the workbook and assigns the active sheet for processing.
        """
        root = Tk()
        root.withdraw()
        self.file = filedialog.askopenfilename(title="Bitte Datei auswählen", initialdir=t.find_path("Sources"),filetypes=(('Excel files', '*.xlsx'), ('All files', '*.*')))
        workbook = load_workbook(self.file)
        self.sheet = workbook.active

    def get_connection(self):
        """
        Establishes a connection to a new SQLite database based on the selected Excel file.

        Verifies that the first three columns of the sheet are "Lektion", "lat", and "art".
        If the columns do not match, an error message is displayed.
        
        If the columns are correct, a new SQLite database is created, named after the Excel file
        (without the extension).
        """

        if (self.file != "") and (self.sheet["A"][0].value != "Lektion" or self.sheet["B"][0].value != "lat" or self.sheet["C"][0].value != "art"):
            print("Datenbank wurde nicht erstellt. Bitte folgende Reihenfolge der Tabelle beachten:\nLektion - lat - art")
        else:
            name_of_new_DB = f'{self.file[:self.file.rfind(".")]}.db'
            if os.path.isfile(t.find_path(name_of_new_DB)):
                os.remove(t.find_path(name_of_new_DB))
            self.connection = sqlite3.connect(name_of_new_DB)

    def create_DB(self):
        """
        Creates tables and inserts vocabulary data from the Excel worksheet into the SQLite database.

        This method creates a table for each lesson and populates it with vocabulary and word type data.
        The database tables are named according to the lesson number, and each table contains
        two columns: "Latein" (Latin word) and "Wortart" (word type).
        """
        pointer = self.connection.cursor()

        num_lecs = []
        for lecs in self.sheet["A"]:
            num_lecs.append(lecs.value)
        num_lecs.pop(0)
        for i in range(1, max(num_lecs)+1):
            pointer.execute(f'CREATE TABLE "Vokabeln_Lektion_{i}" ("Latein" TEXT, "Wortart" TEXT, PRIMARY KEY("Latein") )')

        for value in self.sheet.iter_rows(min_col=1,max_col=3,values_only=True):
            if value[0] == "Lektion":
                pass
            else:
                lektion = value[0]
                vokabel = value[1]
                wortart = value[2]
                pointer.execute(f'INSERT INTO Vokabeln_Lektion_{lektion} VALUES ("{vokabel}", "{wortart}")')
                self.connection.commit()

        self.connection.close()

#main Code for executing
creator = dbCreator()
creator.get_connection()
creator.create_DB()